"""
Monitor de precios y ofertas.
------------------------------
Lee una lista de productos (nombre + URL) desde productos.json,
entra a cada página, extrae precio (y oferta si existe), y agrega
una fila nueva por producto/fecha a un Excel histórico.

Diseñado para correr una vez al día (vía GitHub Actions, Task
Scheduler o cron). Cada corrida AGREGA filas, no las sobreescribe,
así queda un histórico de precios en el tiempo.
"""

import json
import re
import sys
import time
import random
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright

ZONA_HORARIA = ZoneInfo("America/Mexico_City")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/webp,*/*;q=0.8"
    ),
    "Accept-Language": "es-MX,es;q=0.9,en;q=0.8",
}

# Espera entre cada producto para no verse como un bot golpeando el sitio
# muy rápido (algunas tiendas bloquean/limitan peticiones consecutivas).
ESPERA_MIN_SEGUNDOS = 8
ESPERA_MAX_SEGUNDOS = 15

# Si a la primera no se encuentra el precio, cuántas veces reintentar
# (con una espera más larga cada vez) antes de darse por vencido.
REINTENTOS = 3

PRODUCTOS_FILE = Path(__file__).parent / "productos.json"
EXCEL_FILE = Path(__file__).parent / "historial_precios.xlsx"


def extraer_precio(html: str) -> float | None:
    """
    Estrategia genérica: busca el primer patrón tipo '$123.45' o '$1,234.00'
    en el HTML. Funciona bien en sitios VTEX (como TONY) donde el precio
    va en texto plano cerca del título del producto.

    Si en el futuro agregas un sitio donde esto falle, lo ideal es escribir
    una función extractora específica para ese dominio (ver ejemplo abajo
    para VTEX).
    """
    soup = BeautifulSoup(html, "html.parser")
    texto = soup.get_text(" ", strip=True)

    match = re.search(r"\$\s?(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)", texto)
    if match:
        return float(match.group(1).replace(",", ""))
    return None


def extraer_precio_vtex(html: str) -> float | None:
    """
    Extractor específico para tiendas VTEX (como tony.com.mx).
    VTEX suele renderizar el precio como texto plano '$XX.XX' justo
    después del título/SKU del producto en el HTML estático.
    Reutiliza la misma lógica genérica; se deja separado por si en el
    futuro VTEX cambia su forma de renderizar y hay que ajustar solo esto.
    """
    return extraer_precio(html)


def extraer_precio_shopify_meta(html: str) -> float | None:
    """
    Extractor para tiendas Shopify (como grupopapelerogutierrez.com.mx).
    Usa la etiqueta <meta property="product:price:amount" content="XX.XX">
    del HTML, que es mucho más confiable que buscar '$' en el texto visible
    (evita agarrar por error otros montos en la página, como umbrales de
    envío gratis).
    """
    soup = BeautifulSoup(html, "html.parser")
    tag = soup.find("meta", attrs={"property": "product:price:amount"})
    if tag and tag.get("content"):
        try:
            return float(tag["content"])
        except ValueError:
            pass
    # si no encuentra la meta, cae al método genérico como respaldo
    return extraer_precio(html)


# Mapea fragmentos de dominio -> función extractora.
# Agrega aquí una entrada por cada sitio con estructura distinta.
EXTRACTORES = {
    "tony.com.mx": extraer_precio_vtex,
    "grupopapelerogutierrez.com.mx": extraer_precio_shopify_meta,
}


def obtener_extractor(url: str):
    for dominio, func in EXTRACTORES.items():
        if dominio in url:
            return func
    return extraer_precio  # fallback genérico


# Dominios que arman la página con JavaScript (aplicaciones tipo SPA) y por
# lo tanto no se pueden leer con una petición normal (requests) — necesitan
# un navegador simulado (Playwright) que espere a que cargue el contenido.
DOMINIOS_RENDER_JS = {
    "marchand.com.mx",
}


def obtener_precio_render_js(url: str) -> float | None:
    """
    Para sitios que cargan el precio con JavaScript (como marchand.com.mx).
    Abre la página en un navegador headless, espera a que cargue el
    contenido, y busca el precio en el texto ya renderizado.
    """
    with sync_playwright() as p:
        browser = p.chromium.launch()
        try:
            page = browser.new_page(user_agent=HEADERS["User-Agent"])
            page.goto(url, wait_until="networkidle", timeout=30000)
            # Espera extra por si el precio se pinta un poco después de que
            # termina la carga de red.
            page.wait_for_timeout(2000)
            texto = page.inner_text("body")
        finally:
            browser.close()

    match = re.search(r"\$\s?(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)", texto)
    if match:
        return float(match.group(1).replace(",", ""))
    return None


def consultar_producto(nombre: str, url: str) -> dict:
    usa_navegador = any(dominio in url for dominio in DOMINIOS_RENDER_JS)
    extractor = obtener_extractor(url)
    ultimo_error = None

    for intento in range(1, REINTENTOS + 2):  # 1 intento normal + reintentos
        try:
            if usa_navegador:
                precio = obtener_precio_render_js(url)
            else:
                resp = requests.get(url, headers=HEADERS, timeout=20)
                resp.raise_for_status()
                precio = extractor(resp.text)
        except Exception as e:
            ultimo_error = f"error: {e}"
        else:
            if precio is not None:
                return {
                    "fecha": datetime.now(ZONA_HORARIA).strftime("%Y-%m-%d %H:%M"),
                    "producto": nombre,
                    "url": url,
                    "precio": precio,
                    "estado": "ok",
                }
            ultimo_error = "N/A"

        if intento <= REINTENTOS:
            espera = ESPERA_MAX_SEGUNDOS * (intento + 1)  # espera más larga cada reintento
            print(f"  reintento {intento} para '{nombre}' en {espera}s...")
            time.sleep(espera)

    return {
        "fecha": datetime.now(ZONA_HORARIA).strftime("%Y-%m-%d %H:%M"),
        "producto": nombre,
        "url": url,
        "precio": None,
        "estado": ultimo_error,
    }


def guardar_en_excel(filas: list[dict]):
    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial"
        ws.append(["Fecha", "Producto", "Precio", "Estado", "URL"])

    for fila in filas:
        precio_valor = fila["precio"] if fila["precio"] is not None else "N/A"
        ws.append([
            fila["fecha"],
            fila["producto"],
            precio_valor,
            fila["estado"],
            fila["url"],
        ])
        # Si sí se encontró el precio, formatea la celda para que siempre
        # muestre 2 decimales (ej. 50.30 en vez de 50.3).
        if fila["precio"] is not None:
            ws.cell(row=ws.max_row, column=3).number_format = "0.00"

    wb.save(EXCEL_FILE)


def main():
    if not PRODUCTOS_FILE.exists():
        print(f"No encontré {PRODUCTOS_FILE}. Crea ese archivo con tu lista de productos.")
        sys.exit(1)

    productos = json.loads(PRODUCTOS_FILE.read_text(encoding="utf-8"))
    resultados = []

    for i, p in enumerate(productos):
        print(f"Consultando: {p['nombre']}...")
        resultado = consultar_producto(p["nombre"], p["url"])
        resultados.append(resultado)
        print(f"  -> {resultado['estado']} | precio: {resultado['precio']}")

        # Espera entre productos (no después del último) para no golpear
        # el sitio con peticiones seguidas.
        if i < len(productos) - 1:
            espera = random.uniform(ESPERA_MIN_SEGUNDOS, ESPERA_MAX_SEGUNDOS)
            time.sleep(espera)

    guardar_en_excel(resultados)
    print(f"\nListo. {len(resultados)} productos guardados en {EXCEL_FILE}")


if __name__ == "__main__":
    main()
